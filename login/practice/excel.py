#import openpyxl as ex
try:
    from openpyxl import load_workbook

#print(5)
    wb=load_workbook("extest.xlsx")
    exsheet=wb.get_active_sheet()
    
    a=exsheet.cell('b5').value
    
    print(a)
except:
    pass
