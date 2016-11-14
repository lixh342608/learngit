#coding=utf-8
'''
Created on 2016年7月25日

@author: admin
'''
#import openpyxl as ex
try:
    from openpyxl.reader.excel import load_workbook

#print(5)
    wb=load_workbook("extest.xlsx")
    sheetnames=wb.get_sheet_names()
    exsheet=wb.get_sheet_by_name(sheetnames[0])
    
    a=exsheet.cell('b5').value
    
    print(a)
except:
    pass